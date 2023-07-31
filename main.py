# Lista de enteros con codigos de materias que deben estar en la combinacion
# Puede estar vacia
REQUISITOS = []
# Ejemplo
# REQUISITOS = [3628, 3641]
# Todas las combinaciones van a incluir Fisica 1 y Bases de Datos Aplicadas

# Query de SQLite para buscar materias disponibles
# La tabla tiene las columnas
# - Codgio: int
# - Comision: int
# - Turno: ["Mañana", "Tarde", "Noche"]
# - Dia: ["Lunes",
#         'Martes',
#         'Miercoles',
#         'Jueves',
#         'Viernes',
#         'Sabado' ]
EXTRA_QUERY = ''
# Ejemplo
# EXTRA_QUERY = '''
#         ( Turno="Noche"
#         OR (Turno='Tarde' AND NOT Codigo=3628)
#         OR Dia="Sabado" )
#     AND NOT Codigo=3639
#     AND NOT Codigo=3644
#     '''
#   materias a la noche
# o a la tarde (excepto fisica)
# o los sabados
# excepto analisis matematico
#       y gestion de las organizaciones


import random
import sqlite3
from datetime import datetime
import pandas as pd
import hashlib
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import procesar_materias
import procesar_historia_academica
import procesar_oferta_de_materias


DEFAULT_DIR = './files/'
FILE_MATERIAS = 'materias.htm'
FILE_HISTORIA = 'historia.htm'
FILE_OFERTA = 'oferta.htm'
DATABASE = 'materias.db'
TABLE_HISTORIA = 'HISTORIA'
TABLE_OFERTA = 'OFERTA'
TABLE_CORRELATIVAS = 'CORRELATIVAS'
TABLE_MATERIAS = 'MATERIAS'

OUTPUT = f'output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
WORKSHEET = 'Combinaciones'
writer = None

dias = [
    'Lunes',
    'Martes',
    'Miercoles',
    'Jueves',
    'Viernes',
    'Sabado'
]
turnos = [
    'Mañana',
    'Tarde',
    'Noche'
]


def main(file_materias, file_historia, file_oferta):
    global writer
    con = sqlite3.connect(DATABASE)
    cur = con.cursor()

    procesar_materias.load_table(cur, file_materias)
    con.commit()
    procesar_historia_academica.load_table(cur, file_historia)
    con.commit()
    procesar_oferta_de_materias.load_table(cur, file_oferta)
    con.commit()

    ofertas = get_ofertas(cur, EXTRA_QUERY)

    writer = pd.ExcelWriter(OUTPUT)
    generar_combinaciones(cur, ofertas, REQUISITOS)
    # Setear ancho de columnas
    writer.sheets[WORKSHEET].set_column('C:H', 40)
    writer.close()

    set_aligment()


def generar_combinaciones(cur, ofertas, requisitos=None, combinacion=None, i=0, materias=None):
    if requisitos is None:
        requisitos = []
    if materias is None:
        materias = [None for _ in dias]
    if combinacion is None:
        combinacion = {}

    if i >= len(dias):
        if set(requisitos).issubset(materias):
            print_combinacion(cur, combinacion)
        return

    dia = dias[i]

    # Dias que no tienen ninguna materia que no haya sido tomada (o que no tengan materias)
    if set([codigo for codigo, _ in ofertas[dia]]).issubset(materias):
        combinacion[dia] = None
        generar_combinaciones(cur, ofertas, requisitos, combinacion, i + 1, materias)
    else:
        for oferta in ofertas[dia]:
            codigo, _ = oferta
            if codigo in materias:
                continue
            materias[i] = codigo
            combinacion[dia] = oferta
            generar_combinaciones(cur, ofertas, requisitos, combinacion, i + 1, materias)
            materias[i] = None


startrow = 1


# noinspection PyUnresolvedReferences
def print_combinacion(cur, combinacion):
    global startrow
    global writer
    table = {dia: {turno: '' for turno in turnos} for dia in dias}
    for dia in dias:
        if not combinacion[dia]:
            continue
        codigo, comision = combinacion[dia]
        res = cur.execute(f'''
            SELECT M.Nombre, O.Turno 
            FROM {TABLE_OFERTA} O
            INNER JOIN {TABLE_MATERIAS} M ON O.Codigo=M.Codigo 
            WHERE O.Codigo={codigo} AND O.Comision={comision}
        ''')
        nombre, turno = res.fetchone()
        table[dia][turno] = f'{nombre}\n({codigo} - {comision})'

    df = pd.DataFrame(table)

    st = df.style.applymap(generate_style)
    st.to_excel(writer, WORKSHEET, startrow=startrow, startcol=1)
    startrow += 5


def generate_style(string):
    if string:
        # solo quiero hashear la parte del nombre
        # random.seed(hash(string.split('\n')[0]))
        # uso hashlib porque quiero que sea deterministico
        random.seed(hashlib.md5(string.split('\n')[0].encode()).hexdigest())

        # generar color claro
        r = random.randint(120, 235)
        g = random.randint(150, 235)
        b = random.randint(120, 235)

        # convertir a codigo hexadecimal
        color_hex = f'#{r:02x}{g:02x}{b:02x}'
    else:
        color_hex = '#FFFFFF'

    return f'background-color: {color_hex};'


def get_ofertas(cur, condiciones=''):
    crear_vista_disponibles(cur)

    ofertas = {}
    for dia in dias:
        ofertas[dia] = cur.execute(f'''
            SELECT O.Codigo, O.Comision
              FROM {TABLE_OFERTA} O
             WHERE O.Codigo in (SELECT D.Codigo FROM DISPONIBLES D)
               AND O.Dia = "{dia}"
             {'AND (' + condiciones + ')' if condiciones else ''}
        ''').fetchall()

    # res = cur.execute('SELECT Codigo FROM DISPONIBLES')
    # pprint(res.fetchall())

    return ofertas


def crear_vista_disponibles(cur):
    cur.execute(f'''
        CREATE TEMP VIEW APROBADAS AS
        SELECT M.Codigo
        FROM {TABLE_MATERIAS} M
        WHERE M.Nota>=4
    ''')

    cur.execute(f'''
        CREATE TEMP VIEW DISPONIBLES AS
        SELECT M.Codigo
        FROM {TABLE_MATERIAS} M
        WHERE NOT EXISTS (
            SELECT * FROM {TABLE_CORRELATIVAS} C 
            WHERE C.Codigo = M.Codigo AND C.Correlativa NOT IN (SELECT * FROM APROBADAS)
        )
        EXCEPT
        SELECT A.Codigo
        FROM APROBADAS A
    ''')


# Por algun motivo el Styler y el book.add_style no se llevan bien
# Uso openpyxl para asignar wrap text=True en el libro
def set_aligment():
    wb = load_workbook(OUTPUT)
    ws = wb[WORKSHEET]
    # ws['A:H'].alignment = Alignment(wrap_text=True)
    for row in ws['A:H']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.save(OUTPUT)


if __name__ == '__main__':
    main(DEFAULT_DIR + FILE_MATERIAS,
         DEFAULT_DIR + FILE_HISTORIA,
         DEFAULT_DIR + FILE_OFERTA)
